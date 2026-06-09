"""Админ-панель: студенты, гости, приглашения, сотрудники."""
import io
import re
from datetime import timedelta

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from sqlalchemy import func, or_

from ..auth.decorators import admin_required
from ..extensions import db
from ..models import (
    RELATIONS,
    ROLE_ADMIN,
    ROLE_SCANNER,
    STATUS_PRESENT,
    Guest,
    Student,
    User,
)
from ..utils import invite_url

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


@admin_bp.before_request
@login_required
@admin_required
def _guard():
    """Все маршруты админки доступны только администраторам."""
    pass


# ===========================================================================
#  Дашборд
# ===========================================================================
@admin_bp.route("/")
def dashboard():
    students_count = Student.query.count()
    guests_count = Guest.query.count()
    present_count = Guest.query.filter_by(status=STATUS_PRESENT).count()
    invited_count = Guest.query.filter(Guest.token.isnot(None)).count()

    recent = (
        Guest.query.filter_by(status=STATUS_PRESENT)
        .order_by(Guest.checked_in_at.desc())
        .limit(8)
        .all()
    )

    from flask import current_app

    capacity = current_app.config["VENUE_CAPACITY"]

    return render_template(
        "admin/dashboard.html",
        students_count=students_count,
        guests_count=guests_count,
        present_count=present_count,
        invited_count=invited_count,
        capacity=capacity,
        recent=recent,
    )


# ===========================================================================
#  Студенты
# ===========================================================================
@admin_bp.route("/students")
def students():
    q = (request.args.get("q") or "").strip()
    query = Student.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(Student.full_name.ilike(like), Student.group_name.ilike(like))
        )
    items = query.order_by(Student.group_name, Student.full_name).all()
    return render_template("admin/students.html", students=items, q=q)


@admin_bp.route("/students/new", methods=["POST"])
def student_create():
    full_name = (request.form.get("full_name") or "").strip()
    group_name = (request.form.get("group_name") or "").strip()
    specialty = (request.form.get("specialty") or "").strip() or None

    if not full_name:
        flash("Укажите ФИО студента.", "error")
        return redirect(url_for("admin.students"))

    student = Student(full_name=full_name, group_name=group_name, specialty=specialty)
    db.session.add(student)
    db.session.commit()
    flash(f"Студент «{full_name}» добавлен.", "success")
    return redirect(url_for("admin.student_detail", sid=student.id))


@admin_bp.route("/students/<int:sid>")
def student_detail(sid):
    student = db.session.get(Student, sid) or abort(404)
    # Готовим ссылки-приглашения для гостей с токеном
    links = {g.id: invite_url(g.token) for g in student.guests if g.token}
    return render_template(
        "admin/student_detail.html",
        student=student,
        links=links,
        relations=RELATIONS,
    )


@admin_bp.route("/students/<int:sid>/edit", methods=["POST"])
def student_edit(sid):
    student = db.session.get(Student, sid) or abort(404)
    student.full_name = (request.form.get("full_name") or student.full_name).strip()
    student.group_name = (request.form.get("group_name") or "").strip()
    student.specialty = (request.form.get("specialty") or "").strip() or None
    db.session.commit()
    flash("Данные студента обновлены.", "success")
    return redirect(url_for("admin.student_detail", sid=sid))


@admin_bp.route("/students/<int:sid>/delete", methods=["POST"])
def student_delete(sid):
    student = db.session.get(Student, sid) or abort(404)
    db.session.delete(student)
    db.session.commit()
    flash("Студент и связанные гости удалены.", "success")
    return redirect(url_for("admin.students"))


# ===========================================================================
#  Гости (родители / родственники)
# ===========================================================================
@admin_bp.route("/students/<int:sid>/guests", methods=["POST"])
def guest_create(sid):
    from flask import current_app

    student = db.session.get(Student, sid) or abort(404)

    limit = current_app.config["MAX_GUESTS_PER_STUDENT"]
    if student.guests_total >= limit:
        flash(
            f"Достигнут лимит гостей на студента ({limit}). "
            "Удалите кого-то или измените лимит в настройках.",
            "error",
        )
        return redirect(url_for("admin.student_detail", sid=sid))

    full_name = (request.form.get("full_name") or "").strip()
    relation = (request.form.get("relation") or "Родственник").strip()
    phone = (request.form.get("phone") or "").strip() or None

    if not full_name:
        flash("Укажите ФИО гостя.", "error")
        return redirect(url_for("admin.student_detail", sid=sid))

    guest = Guest(
        student_id=student.id,
        full_name=full_name,
        relation=relation,
        phone=phone,
    )
    db.session.add(guest)
    db.session.flush()

    # Защита от гонки/двойного сабмита: повторно считаем в той же транзакции.
    count = (
        db.session.query(func.count(Guest.id))
        .filter_by(student_id=student.id)
        .scalar()
    )
    if count > limit:
        db.session.rollback()
        flash(f"Достигнут лимит гостей на студента ({limit}).", "error")
        return redirect(url_for("admin.student_detail", sid=sid))

    db.session.commit()
    flash(f"Гость «{full_name}» добавлен.", "success")
    return redirect(url_for("admin.student_detail", sid=sid) + f"#guest-{guest.id}")


@admin_bp.route("/guests/<int:gid>/edit", methods=["POST"])
def guest_edit(gid):
    guest = db.session.get(Guest, gid) or abort(404)
    guest.full_name = (request.form.get("full_name") or guest.full_name).strip()
    guest.relation = (request.form.get("relation") or guest.relation).strip()
    guest.phone = (request.form.get("phone") or "").strip() or None
    db.session.commit()
    flash("Данные гостя обновлены.", "success")
    return redirect(
        url_for("admin.student_detail", sid=guest.student_id) + f"#guest-{guest.id}"
    )


@admin_bp.route("/guests/<int:gid>/invite", methods=["POST"])
def guest_invite(gid):
    guest = db.session.get(Guest, gid) or abort(404)
    guest.generate_invite()
    db.session.commit()
    flash("Пригласительное сгенерировано. Ссылку можно скопировать ниже.", "success")
    return redirect(
        url_for("admin.student_detail", sid=guest.student_id) + f"#guest-{guest.id}"
    )


@admin_bp.route("/guests/<int:gid>/reset", methods=["POST"])
def guest_reset(gid):
    guest = db.session.get(Guest, gid) or abort(404)
    guest.reset_presence()
    db.session.commit()
    flash("Отметка о присутствии снята.", "success")
    return redirect(
        url_for("admin.student_detail", sid=guest.student_id) + f"#guest-{guest.id}"
    )


@admin_bp.route("/guests/<int:gid>/delete", methods=["POST"])
def guest_delete(gid):
    guest = db.session.get(Guest, gid) or abort(404)
    sid = guest.student_id
    db.session.delete(guest)
    db.session.commit()
    flash("Гость удалён.", "success")
    return redirect(url_for("admin.student_detail", sid=sid))


@admin_bp.route("/guests")
def guests_overview():
    """Общий список гостей — удобно во время мероприятия."""
    status = request.args.get("status") or ""
    q = (request.args.get("q") or "").strip()

    query = Guest.query.join(Student)
    if status:
        query = query.filter(Guest.status == status)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Guest.full_name.ilike(like),
                Student.full_name.ilike(like),
                Student.group_name.ilike(like),
            )
        )
    items = query.order_by(Guest.status.desc(), Guest.full_name).all()
    return render_template(
        "admin/guests.html", guests=items, status=status, q=q
    )


# ===========================================================================
#  Сотрудники (учётные записи)
# ===========================================================================
@admin_bp.route("/users")
def users():
    items = User.query.order_by(User.role, User.username).all()
    return render_template("admin/users.html", users=items)


@admin_bp.route("/users/new", methods=["POST"])
def user_create():
    username = (request.form.get("username") or "").strip()
    full_name = (request.form.get("full_name") or "").strip()
    role = request.form.get("role") or ROLE_SCANNER
    password = request.form.get("password") or ""

    if role not in (ROLE_ADMIN, ROLE_SCANNER):
        role = ROLE_SCANNER

    if not username or not password:
        flash("Логин и пароль обязательны.", "error")
        return redirect(url_for("admin.users"))
    if len(password) < 8:
        flash("Пароль должен быть не короче 8 символов.", "error")
        return redirect(url_for("admin.users"))
    if User.query.filter_by(username=username).first():
        flash("Пользователь с таким логином уже существует.", "error")
        return redirect(url_for("admin.users"))

    user = User(username=username, full_name=full_name, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash(f"Пользователь «{username}» создан.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:uid>/toggle", methods=["POST"])
def user_toggle(uid):
    user = db.session.get(User, uid) or abort(404)
    if user.id == current_user.id:
        flash("Нельзя отключить собственную учётную запись.", "error")
        return redirect(url_for("admin.users"))
    user.active = not user.active
    db.session.commit()
    flash("Статус учётной записи обновлён.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:uid>/delete", methods=["POST"])
def user_delete(uid):
    user = db.session.get(User, uid) or abort(404)
    if user.id == current_user.id:
        flash("Нельзя удалить собственную учётную запись.", "error")
        return redirect(url_for("admin.users"))
    db.session.delete(user)
    db.session.commit()
    flash("Пользователь удалён.", "success")
    return redirect(url_for("admin.users"))


# ===========================================================================
#  Экспорт по группам (Excel со ссылками-приглашениями)
# ===========================================================================
def _distinct_groups() -> list[str]:
    rows = (
        db.session.query(Student.group_name)
        .distinct()
        .order_by(Student.group_name)
        .all()
    )
    return [r[0] for r in rows if r[0] is not None]


def _fmt_date(value) -> str:
    if not value:
        return ""
    offset = int(current_app.config.get("DISPLAY_TZ_OFFSET", 5))
    return (value + timedelta(hours=offset)).strftime("%d.%m.%Y")


def _safe_sheet_title(name: str, used: set) -> str:
    title = re.sub(r"[:\\/?*\[\]]", " ", name or "Группа").strip()[:31] or "Группа"
    base, candidate, i = title, title, 1
    while candidate in used:
        i += 1
        suffix = f" ({i})"
        candidate = base[: 31 - len(suffix)] + suffix
    used.add(candidate)
    return candidate


def _build_export(groups: list[str]) -> io.BytesIO:
    """Книга Excel: по листу на группу, с готовыми ссылками-приглашениями."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = [
        "№", "ФИО студента", "Группа", "ФИО гостя", "Кем приходится",
        "Ссылка-приглашение", "Действует до", "Статус",
    ]
    widths = [5, 34, 16, 30, 14, 52, 14, 14]

    wb = Workbook()
    wb.remove(wb.active)
    used: set = set()
    header_fill = PatternFill("solid", fgColor="0A2E5C")
    header_font = Font(color="FFFFFF", bold=True)

    for group in groups:
        students = (
            Student.query.filter_by(group_name=group)
            .order_by(Student.full_name)
            .all()
        )
        ws = wb.create_sheet(_safe_sheet_title(group, used))
        ws.append(headers)
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        n = 0
        for st in students:
            if st.guests:
                for g in st.guests:
                    n += 1
                    link = invite_url(g.token) if g.token else ""
                    ws.append([
                        n, st.full_name, st.group_name, g.full_name, g.relation,
                        link, _fmt_date(g.expires_at), g.status_label,
                    ])
                    if link:
                        c = ws.cell(row=ws.max_row, column=6)
                        c.hyperlink = link
                        c.font = Font(color="1F5FAE", underline="single")
            else:
                n += 1
                ws.append([n, st.full_name, st.group_name, "", "", "", "", "нет гостя"])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@admin_bp.route("/export")
def export_page():
    groups = _distinct_groups()
    info = []
    for g in groups:
        students = Student.query.filter_by(group_name=g).all()
        guests = sum(len(s.guests) for s in students)
        with_link = sum(1 for s in students for x in s.guests if x.token)
        info.append({
            "group": g,
            "students": len(students),
            "guests": guests,
            "links": with_link,
        })
    return render_template("admin/export.html", groups=info)


@admin_bp.route("/export/all.xlsx")
def export_all():
    groups = _distinct_groups()
    if not groups:
        flash("Нет данных для экспорта.", "error")
        return redirect(url_for("admin.export_page"))
    bio = _build_export(groups)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Приглашения по группам.xlsx",
    )


@admin_bp.route("/export/group.xlsx")
def export_group():
    group = (request.args.get("group") or "").strip()
    if not group or not Student.query.filter_by(group_name=group).first():
        abort(404)
    bio = _build_export([group])
    safe = re.sub(r"[\\/:*?\"<>|]", "_", group)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Приглашения {safe}.xlsx",
    )
